import pandas as pd
import os
import csv
import time
import logging
import traceback
import shutil
import json
from datetime import datetime
from dask.distributed import Client, LocalCluster, fire_and_forget
from zipfile import ZipFile, ZIP_DEFLATED
from openpyxl.utils import get_column_letter

# connect_to_mongoDB, connect_to_mongo_gridfs, reduced_file, api_search_masses, api_search_formulas,
from .utilities import *

from .heatmap import *
from .cv_scatterplot import *

from . import task_functions as task_fun
from .WebApp_plotter import WebApp_plotter
import io

logger = logging.getLogger("nta_app.ms1")

# import seaborn as sns
try:
    import seaborn as sns
except ModuleNotFoundError:
    logger.error("Seaborn is not installed. Please run 'pip install seaborn' to install it.")

# set this to true to run locally without test (for debug purposes)
NO_DASK = False


def run_nta_dask(
    parameters,
    input_dfs,
    tracer_df=None,
    run_sequence_pos_df=None,
    run_sequence_neg_df=None,
    jobid="00000000",
    verbose=True,
):
    in_docker = os.environ.get("IN_DOCKER") != "False"
    mongo_address = os.environ.get("MONGO_SERVER")

    if NO_DASK:
        run_nta(
            parameters,
            input_dfs,
            tracer_df,
            run_sequence_pos_df,
            run_sequence_neg_df,
            mongo_address,
            jobid,
            verbose,
            in_docker=in_docker,
        )
        return
    if not in_docker:
        logger.info("Running in local development mode.")
        logger.info("Detected OS is {}".format(os.environ.get("SYSTEM_NAME")))
        local_cluster = LocalCluster(processes=False)
        dask_client = Client(local_cluster)
    else:
        dask_scheduler = os.environ.get("DASK_SCHEDULER")
        logger.info("Running in docker environment. Dask Scheduler: {}".format(dask_scheduler))
        dask_client = Client(dask_scheduler)

    input_dfs_size = len(input_dfs)
    logger.info("Before scatter input_dfs_size: {}".format(input_dfs_size))
    dask_input_dfs = dask_client.scatter(input_dfs)
    dask_input_dfs_size = len(dask_input_dfs)
    logger.info("After scatter dask_input_dfs_size: {}".format(dask_input_dfs_size))

    logger.warning("Submitting Nta Dask task")
    task = dask_client.submit(
        run_nta,
        parameters,
        dask_input_dfs,
        tracer_df,
        run_sequence_pos_df,
        run_sequence_neg_df,
        mongo_address,
        jobid,
        verbose,
        in_docker=in_docker,
    )
    fire_and_forget(task)


def run_nta(
    parameters,
    input_dfs,
    tracer_df=None,
    run_sequence_pos_df=None,
    run_sequence_neg_df=None,
    mongo_address=None,
    jobid="00000000",
    verbose=True,
    in_docker=True,
):
    nta_run = NtaRun(
        parameters,
        input_dfs,
        tracer_df,
        run_sequence_pos_df,
        run_sequence_neg_df,
        mongo_address,
        jobid,
        verbose,
        in_docker=in_docker,
    )
    try:
        nta_run.execute()
    except Exception as e:
        trace = traceback.format_exc()
        logger.info(trace)
        fail_step = nta_run.get_step()
        nta_run.set_status("Failed on step: " + fail_step)
        error = repr(e)
        nta_run.set_except_message(error)
        raise e
    return True


class NtaRun:
    def __init__(
        self,
        parameters=None,
        input_dfs=None,
        tracer_df=None,
        run_sequence_pos_df=None,
        run_sequence_neg_df=None,
        mongo_address=None,
        jobid="00000000",
        verbose=True,
        in_docker=True,
    ):
        logger.info(f"\n============= Job ID: {jobid}")
        logger.info("[Job ID: {}] Initializing NtaRun Task".format(jobid))
        logger.info("parameters= {}".format(parameters))
        self.parameters = parameters
        self.tracer_df = tracer_df
        self.tracer_dfs_out = None
        self.run_sequence_pos_df = run_sequence_pos_df
        self.run_sequence_neg_df = run_sequence_neg_df
        self.dfs = input_dfs
        self.dfs_flagged = None  # DFs that will retain occurrences failing CV values
        self.docs = None
        self.doc_combined = None
        self.df_combined = None
        self.df_flagged_combined = None
        self.pass_through = None
        self.mpp_ready = None
        self.mpp_ready_flagged = None
        self.search_results = None
        self.search = None
        self.jobid = jobid
        self.verbose = verbose
        self.in_docker = in_docker
        self.mongo_address = mongo_address
        self.mongo = connect_to_mongoDB(self.mongo_address)
        self.gridfs = connect_to_mongo_gridfs(self.mongo_address)
        self.base_dir = os.path.abspath(os.path.join(os.path.abspath(__file__), "../../.."))
        self.data_map = {}
        self.tracer_map = {}
        self.occurrence_heatmap_map = {}
        self.cv_scatterplot_map = {}
        # self.data_dir = os.path.join(self.base_dir, 'data', self.jobid)
        # self.new_download_dir = os.path.join(self.data_dir, "new")
        self.step = "Started"  # tracks the current step (for fail messages)
        # os.mkdir(self.data_dir)
        # os.mkdir(self.new_download_dir)
        self.tracer_plots_out = []
        self.occurrence_heatmaps_out = []
        self.cv_scatterplots_out = []

        # NTAW-594
        self.all_headers, self.blank_headers, self.sample_headers = task_fun.get_sample_and_blank_headers(self.dfs)
        logger.info(f"all headers: {self.all_headers}")
        logger.info(f"blank headers: {self.blank_headers}")
        logger.info(f"sample headers: {self.sample_headers}")

    def execute(self):
        self.step = "Check for existence of required columns"
        # 1a: check existence of "Ionization mode" column
        self.check_existence_of_ionization_mode_column(self.dfs)
        # 1b: check existence of 'mass column'
        self.check_existence_of_mass_column(self.dfs)
        # 1c: check for alternate spellings of 'Retention_Time' column
        self.check_retention_time_column(self.dfs)
        # 1d: sort dataframe columns alphabetically
        self.dfs = [df.reindex(sorted(df.columns), axis=1) if df is not None else None for df in self.dfs]
        # 1e: create a status in mongo
        self.set_status("Processing", create=True)
        # 1f: create an analysis_parameters sheet
        self.create_analysis_parameters_sheet()
        # 1g: create run sequence sheets
        self.create_run_sequence_sheets()
        # 2: assign ids, separate passthrough cols, filter void volume, and flag duplicates
        self.step = "Flagging duplicates"
        self.assign_id()
        self.pass_through_cols()
        self.filter_void_volume(float(self.parameters["minimum_rt"][1]))  # throw out features below this (void volume)
        self.filter_duplicates()
        if self.verbose:
            logger.info("Flagged duplicates.")
            logger.info("dfs.size(): {}".format(len(self.dfs)))
            if self.dfs[0] is not None:
                logger.info("POS df length: {}".format(len(self.dfs[0])))
                logger.info("POS df columns: {}".format(self.dfs[0].columns))
            if self.dfs[1] is not None:
                logger.info("NEG df length: {}".format(len(self.dfs[1])))
                logger.info("NEG df columns: {}".format(self.dfs[1].columns))

        # 3a: statistics
        self.step = "Calculating statistics"
        self.calc_statistics()
        if self.verbose:
            logger.info("Calculated statistics.")
            if self.dfs[0] is not None:
                logger.info("POS df length: {}".format(len(self.dfs[0])))
            if self.dfs[1] is not None:
                logger.info("NEG df length: {}".format(len(self.dfs[1])))

        # 3b: Occurrence heatmap
        self.step = "Create heatmap"
        self.store_heatmap()

        # 4a: check tracers (optional)
        self.step = "Checking tracers"
        if self.verbose:
            logger.info("Checking tracers: DF check for debugging.")
            logger.info("dfs.size(): {}".format(len(self.dfs)))
            if self.dfs[0] is not None:
                logger.info("POS df length: {}".format(len(self.dfs[0])))
                logger.info("POS df columns: {}".format(self.dfs[0].columns.tolist()))
            if self.dfs[1] is not None:
                logger.info("NEG df length: {}".format(len(self.dfs[1])))
                logger.info("NEG df columns: {}".format(self.dfs[1].columns.tolist()))
        self.check_tracers()
        if self.verbose:
            logger.info("Checked tracers.")

        # 4b: CV Scatterplot
        self.step = "Create scatterplot"
        self.store_scatterplots()

        # 5a: clean features
        self.step = "Cleaning features"
        self.clean_features()

        if self.verbose:
            logger.info("Cleaned features.")
            if self.dfs[0] is not None:
                logger.info("POS df length: {}".format(len(self.dfs[0])))
            if self.dfs[1] is not None:
                logger.info("NEG df length: {}".format(len(self.dfs[1])))

        # 5b: Merge detection count columns onto tracers for export
        self.step = "Merge detection counts onto tracers"
        self.merge_columns_onto_tracers()

        if self.verbose:
            logger.info("Created flags.")
            if self.dfs[0] is not None:
                logger.info("POS df length: {}".format(len(self.dfs[0])))
                # NTAW-94 set flag for dashboard search
                self.dfs[0]["For_Dashboard_Search"] = "1"
                self.dfs_flagged[0]["For_Dashboard_Search"] = "1"
            if self.dfs[1] is not None:
                logger.info("NEG df length: {}".format(len(self.dfs[1])))
                # NTAW-94 set flag for dashboard search
                self.dfs[1]["For_Dashboard_Search"] = "1"
                self.dfs_flagged[1]["For_Dashboard_Search"] = "1"

        # 6: combine modes
        self.step = "Combining modes"
        self.combine_modes()
        if self.verbose:
            logger.info("Combined modes.")
            logger.info("combined df length: {}".format(len(self.df_combined)))

        # 7: search dashboard
        if self.parameters["search_dsstox"][1] == "yes":
            self.step = "Searching dsstox database"
            self.perform_dashboard_search()
            if self.parameters["search_hcd"][1] == "yes":
                self.step = "Searching Cheminformatics Hazard Module database"
                self.perform_hcd_search()

        # 8: Store excel data to MongoDB
        self.step = "Storing data"
        self.save_excel_to_mongo()

        # 9: set status to completed
        self.step = "Displaying results"
        self.set_status("Completed")
        logger.warning("MS1 job {}: Processing complete.".format(self.jobid))

    def check_existence_of_ionization_mode_column(self, input_dfs):
        """
        Check and ensure the existence of the 'Ionization_Mode' column in a list of DataFrames.

        This function iterates through a list of DataFrames, typically representing positive and negative ionization modes,
        and checks if the 'Ionization_Mode' column is present. If not found, it adds the column to the DataFrame with
        predefined values based on the mode.

        Args:
            self: The instance of the class (typically associated with object-oriented programming).
            input_dfs (list of pandas.DataFrame): A list of pandas DataFrames to check and modify.
        Returns:
            None: This function operates in place and modifies the input DataFrames.
        Example:
            input_dfs = [positive_mode_df, negative_mode_df]
            checker = IonizationModeChecker()
            checker.check_existence_of_ionization_mode_column(input_dfs)
            # The 'Ionization_Mode' column will be added to DataFrames if missing, with values 'Esi+' for positive mode
            # and 'Esi-' for negative mode.
        """
        # the zeroth element of input_dfs is the positive mode dataframe
        ionizationMode = "Esi+"
        # Iterate through list of dfs
        for df in input_dfs:
            if df is not None:
                if "Ionization_Mode" not in df.columns:
                    # create a new column with the header of "Ionization_Mode" and values of ionizationMode
                    df["Ionization_Mode"] = ionizationMode
            # the first element of input_dfs is the negative mode dataframe
            ionizationMode = "Esi-"
        return

    def check_existence_of_mass_column(self, input_dfs):
        """
        Check the existence of a 'Mass' or 'm/z' column in input dataframes and handle it accordingly.
        This function checks each dataframe in the input list for the presence of a 'Mass' or 'm/z' column. If either of these columns is found, it takes appropriate action based on the ionization mode. If neither column exists, it raises a ValueError.

        Args:
            input_dfs (list of pandas DataFrames): A list of dataframes to check.
        Note:
            - If 'Mass' already exists, no action is taken.
            - If 'm/z' exists, a 'Mass' column is created by subtracting 1.0073 for "Esi+" mode and adding 1.0073 for "Esi-" mode.
        Raises:
            ValueError: If neither 'Mass' nor 'm/z' columns are present in the input dataframe.
        Returns:
            None
        """
        ionizationMode = "Esi+"

        # Iterate through two dataframes (assumes there are two in the input)
        for i in range(0, 2):
            df = input_dfs[i]

            # If it's the second iteration, change the ionization mode to "Esi-"
            if i == 1:
                ionizationMode = "Esi-"

            if df is not None:
                if "Mass" in df.columns:
                    pass  # 'Mass' column already exists, no action needed
                elif "m/z" in df.columns:
                    if ionizationMode == "Esi+":
                        df["Mass"] = df["m/z"] - 1.0073
                    elif ionizationMode == "Esi-":
                        df["Mass"] = df["m/z"] + 1.0073
                else:
                    raise ValueError("Either Mass or m/z column must be in the input file. (Check spelling!)")

        return

    def check_retention_time_column(self, input_dfs):
        """
        Check for the existence of alternate spellings of 'Retention_Time' column in input dataframes and rename to "Retention_Time".

        Args:
            input_dfs (list of pandas DataFrames): A list of dataframes to check.
        Raises:
            ValueError: If 'Retention_Time' column is not present in the input dataframe.
        Returns:
            None
        """
        # Iterate through list of dfs
        for df in input_dfs:
            if df is not None:
                # Check to see if there is not the expected spelling of "Retention_Time" column
                # if 'Retention_Time' not in df.columns:
                # replace alternative capitalizations
                df.rename(
                    columns={
                        "Retention_time": "Retention_Time",
                        "RETENTION_TIME": "Retention_Time",
                    },
                    inplace=True,
                )
                # replace rt/RT
                df.rename(
                    columns={"rt": "Retention_Time", "RT": "Retention_Time"},
                    inplace=True,
                )
                # replace "Ret. Time" (SCIEX data)
                df.rename(columns={"Ret._Time": "Retention_Time"}, inplace=True)
                if "Retention_Time" not in df.columns:
                    raise ValueError("Retention_Time column must be in the input file. (Check spelling!)")

        return

    def create_analysis_parameters_sheet(self):
        """
        Create a dataframe to store analysis parameters in, assign parameters, then save
        as the "Analysis Parameters" sheet in the self.data_map dictionary.

        Args:
            None
        Notes:
            In a future version, we would like to add a "Version" key to be printed.
        Returns:
            None
        """
        # create a dataframe to store analysis parameters
        columns = ["Parameter", "Value"]
        df_analysis_parameters = pd.DataFrame(columns=columns)

        # loop through keys in self.parameters and log them
        for key in self.parameters:
            logger.info("key: {}".format(key))
            label = self.parameters[key][0]
            value = self.parameters[key][1]
            df_analysis_parameters.loc[len(df_analysis_parameters)] = [label, value]

        # add the dataframe to the data_map with the sheet name of 'Analysis Parameters'
        self.data_map["Analysis Parameters"] = df_analysis_parameters

        return

    def create_run_sequence_sheets(self):
        """
        If there are run sequence files submitted, create a sheet for each mode of run sequence file. This is for the AMOS visualizations to be able to grab the sequence information for
        run seequence plots

        Args:
            None
        Notes:
            In a future version, we would like to directly pass this information to AMOS separately rather than storing it in the results file where it increases the complexity of the results
        Returns:
            None
        """

        if self.run_sequence_pos_df is not None:
            self.data_map["Run Sequence (pos)"] = self.run_sequence_pos_df
        if self.run_sequence_neg_df is not None:
            self.data_map["Run Sequence (neg)"] = self.run_sequence_neg_df

        return

    def set_status(self, status, create=False):
        """
        Accepts a string (e.g., "Processing", or "Step Completed") and, if create is TRUE,
        post status to the logger with the Job ID and timestamp.

        Args:
            status (string)
            create (Boolean)
        Notes:
            In a future version, we would like to display status on the UI
        Returns:
            None
        """
        posts = self.mongo.posts
        # Get time stamp
        time_stamp = datetime.utcnow()
        # Generate ID with status
        post_id = self.jobid + "_" + "status"
        # If create, post update
        if create:
            posts.update_one(
                {"_id": post_id},
                {
                    "$set": {
                        "_id": post_id,
                        "date": time_stamp,
                        "status": status,
                        "error_info": "",
                    }
                },
                upsert=True,
            )
        else:
            posts.update_one(
                {"_id": post_id},
                {"$set": {"_id": post_id, "status": status}},
                upsert=True,
            )

    def set_except_message(self, e):
        """
        Accepts a string with error/exception information and then
        posts string to the logger with the Job ID and timestamp.

        Args:
            e (string)
        Notes:
            We are working to provide more detailed exceptions
        Returns:
            None
        """
        posts = self.mongo.posts
        # Get timestamp
        time_stamp = datetime.utcnow()
        # Generate ID with status
        post_id = self.jobid + "_" + "status"
        # Post string update
        posts.update_one(
            {"_id": post_id},
            {"$set": {"_id": post_id, "date": time_stamp, "error_info": e}},
            upsert=True,
        )

    def get_step(self):
        """
        Called if "try/ nta_run.execute()" (lines 112-113) fails; returns current
        step of the execute() function for diagnosing error.

        Args:
            None
        Returns:
            self.step (string)
        """
        # Return step string
        return self.step

    def assign_id(self):
        """
        Accesses self.dfs (list of dataframes) and calls task_fun.assign_feature_id()
        on each dataframe. Inserts column "Feature_ID" in the front of each dataframe
        with a unique numeric identifier for each row.

        Args:
            None
        Returns:
            None
        """
        # If both df exists, submit both to assign_feature_id() function, update start
        # kwarg by length of first df
        if self.dfs[0] is not None and self.dfs[1] is not None:
            self.dfs[0] = task_fun.assign_feature_id(self.dfs[0])
            self.dfs[1] = task_fun.assign_feature_id(self.dfs[1], start=len(self.dfs[0].index) + 1)
        # Else, submit single df to assign_feature_id() function
        elif self.dfs[0] is not None:
            self.dfs[0] = task_fun.assign_feature_id(self.dfs[0])
        else:
            self.dfs[1] = task_fun.assign_feature_id(self.dfs[1])
        return

    def pass_through_cols(self):
        """
        Accesses self.dfs (list of dataframes) and calls task_fun.passthrucol()
        on each dataframe. Identifies columns in dataframes that aren't needed for
        downstream functions; removes them from self.dfs and stores them in class
        object self.pass_through. These columns will get appended to the results.

        Args:
            None
        Notes:
            There is probably a more elegant solution than effectively running the
            task_fun.passthrucol() function twice, but I would run into errors if only
            one dataframe in self.dfs was present.
        Returns:
            None
        """
        # Submit dfs and all_headers to passthrucol() function, store outputs separately
        self.pass_through = [
            task_fun.passthrucol(df, self.all_headers)[0] if df is not None else None for df in self.dfs
        ]
        self.dfs = [task_fun.passthrucol(df, self.all_headers)[1] if df is not None else None for df in self.dfs]
        return

    def filter_void_volume(self, min_rt):
        """
        Accesses self.dfs (list of dataframes) and self.parameters["minimum_rt"][1]
        then removes all rows with a value below "minimum_rt" in the "Retention_Time"
        column.

        Args:
            min_rt (float, user-submitted value with default of 0.0)
        Returns:
            None
        """
        # Iterate through dfs, removing rows where "Retention_Time" is below min_rt threshold
        self.dfs = [df.loc[df["Retention_Time"] > min_rt].copy() if df is not None else None for df in self.dfs]
        return

    def filter_duplicates(self):
        """
        Accesses self.dfs (list of dataframes), self.parameters["mass_accuracy_units"][1],
        self.parameters["mass_accuracy"][1], and self.parameters["rt_accuracy"][1].
        The task_fun.duplicates() function is called on each dataframe, with mass_accuracy,
        rt_accuracy, and ppm kwargs passed. The function identifies duplicate features
        within the passed tolerances, flagged in a new "Duplicates Features?" column.

        Args:
            None
        Returns:
            None
        """
        # Get ppm, mass_accuracy, and rt_accuracy parameters
        ppm = self.parameters["mass_accuracy_units"][1] == "ppm"
        mass_accuracy = float(self.parameters["mass_accuracy"][1])
        rt_accuracy = float(self.parameters["rt_accuracy"][1])
        # Perform duplicate flagging functions
        self.dfs = [
            task_fun.duplicates(df, mass_accuracy, rt_accuracy, ppm, self.blank_headers, self.sample_headers)
            if df is not None
            else None
            for df in self.dfs
        ]
        return

    def calc_statistics(self):
        """
        Accesses self.dfs (list of dataframes), self.parameters["mass_accuracy_units"][1],
        self.parameters["mass_accuracy"][1], self.parameters["rt_accuracy"][1], and self.parameters["mrl_std_multiplier"][1].
        The task_fun.chunk_stats() function is called on each dataframe, with mrl_std_multiplier passed. The function calculates
        mean, median, std_dev, CV, and replication % for each chemical feature across sample replicates. If adducts are
        selected by the user, the task_fun.adduct_identifier() function is called on each dataframe. Finally, the
        task_fun.column_sort_DFS() function is called to prepare the dataframes for output in the excel file.

        Args:
            None
        Returns:
            None
        """
        # Get ppm, mass_accuracy, rt_accuracy, and mrl_multiplier parameters
        ppm = self.parameters["mass_accuracy_units"][1] == "ppm"
        mass_accuracy = float(self.parameters["mass_accuracy"][1])
        rt_accuracy = float(self.parameters["rt_accuracy"][1])
        mrl_multiplier = float(self.parameters["mrl_std_multiplier"][1])
        # NTAW-509 Get miminum blank detection percentage
        min_blank_detection_percentage = float(self.parameters["min_replicate_hits_blanks"][1])
        # Iterate through dfs, calling chunk_stats() function
        # NTAW-49: Raises custom ValueError if blank columns are improperly named in the input dataframes
        try:
            self.dfs = [
                task_fun.chunk_stats(
                    df,
                    min_blank_detection_percentage,
                    self.blank_headers,
                    self.sample_headers,
                    mrl_multiplier=mrl_multiplier,
                )
                if df is not None
                else None
                for df in self.dfs
            ]
        except IndexError:
            raise ValueError(
                "Blank samples not found. Blanks must have one of the following text strings present: ['mb', 'Mb','mB', 'MB', 'blank', 'Blank', 'BLANK']"
            )
        # Get positive adducts, print to logger
        pos_adducts_selected = self.parameters["pos_adducts"][1]
        logger.info("pos adducts list: {}".format(self.parameters["pos_adducts"]))
        # Get negative adducts, print to logger
        neg_adducts_selected = self.parameters["neg_adducts"][1]
        logger.info("neg adducts list: {}".format(self.parameters["neg_adducts"]))
        # Get neutral losses, print to logger
        neutral_losses_selected = self.parameters["neutral_losses"][1]
        logger.info("neutral losses list: {}".format(self.parameters["neutral_losses"]))
        # Package adduct lists together into list of lists
        adduct_selections = [pos_adducts_selected, neg_adducts_selected, neutral_losses_selected]
        # Check if any adducts are selected
        to_run = False
        for li in adduct_selections:
            if len(li) > 0:
                to_run = True
        # Check if any adducts selected; if so, perform adduct identification functions
        if to_run:
            if self.dfs[0] is not None and self.dfs[1] is not None:
                self.dfs[0] = task_fun.adduct_identifier(
                    self.dfs[0], adduct_selections, mass_accuracy, rt_accuracy, ppm, ionization="positive"
                )
                self.dfs[1] = task_fun.adduct_identifier(
                    self.dfs[1], adduct_selections, mass_accuracy, rt_accuracy, ppm, ionization="negative"
                )
            elif self.dfs[0] is not None:
                self.dfs[0] = task_fun.adduct_identifier(
                    self.dfs[0], adduct_selections, mass_accuracy, rt_accuracy, ppm, ionization="positive"
                )
            else:
                self.dfs[1] = task_fun.adduct_identifier(
                    self.dfs[1], adduct_selections, mass_accuracy, rt_accuracy, ppm, ionization="negative"
                )
        # sort dataframe columns for data_feature_statistics (DFS) output
        if self.dfs[0] is not None and self.dfs[1] is not None:
            self.data_map["All Detection Statistics (Pos)"] = task_fun.column_sort_DFS(
                self.dfs[0], self.pass_through[0], self.all_headers
            )
            self.data_map["All Detection Statistics (Neg)"] = task_fun.column_sort_DFS(
                self.dfs[1], self.pass_through[1], self.all_headers
            )
        elif self.dfs[0] is not None:
            self.data_map["All Detection Statistics (Pos)"] = task_fun.column_sort_DFS(
                self.dfs[0], self.pass_through[0], self.all_headers
            )
        else:
            self.data_map["All Detection Statistics (Neg)"] = task_fun.column_sort_DFS(
                self.dfs[1], self.pass_through[1], self.all_headers
            )
        return

    def store_scatterplots(self):
        # Store in class variable
        self.cv_scatterplots_out.append(cv_scatterplot(parameters=self.parameters, data_map=self.data_map))
        # Map to outputs
        self.cv_scatterplot_map["cv_scatterplot"] = self.cv_scatterplots_out[0]
        project_name = self.parameters["project_name"][1]
        self.gridfs.put(
            "&&".join(self.cv_scatterplot_map.keys()),
            _id=self.jobid + "_cv_scatterplots",
            encoding="utf-8",
            project_name=project_name,
        )
        # Save to MongoDB
        self.mongo_save(self.cv_scatterplot_map["cv_scatterplot"], step="cv_scatterplot")

    def store_heatmap(self):
        # Store in class variable
        self.occurrence_heatmaps_out.append(occurrence_heatmap(parameters=self.parameters, data_map=self.data_map))
        # Map to outputs
        self.occurrence_heatmap_map["occurrence_heatmap"] = self.occurrence_heatmaps_out[0]
        project_name = self.parameters["project_name"][1]
        self.gridfs.put(
            "&&".join(self.occurrence_heatmap_map.keys()),
            _id=self.jobid + "_occurrence_heatmaps",
            encoding="utf-8",
            project_name=project_name,
        )
        # Save to MongoDB
        self.mongo_save(self.occurrence_heatmap_map["occurrence_heatmap"], step="occurrence_heatmap")

    def check_tracers(self):
        """
        Check for tracers. If present, access user input parameters for mass accuracy units,
        mass accuracy, y-axis scaling (run sequence plots), and trendline toggling (yes/no).
        Access self.dfs and pass each dataframe to task_fun.check_feature_tracers() with
        other parameters as kwargs. Call fromat_tracer_file() function imported from
        utilities.py, then instantiate run sequence plots with the WebApp_plotter. Access
        pass-through columns (if present), and call task_fun.column_sort_TSR() function
        to format the tracer outputs for the excel file.

        Args:
            None
        Returns:
            None
        """
        # If no tracer file present, skip step
        if self.tracer_df is None:
            logger.info("No tracer file, skipping this step.")
            return
        if self.verbose:
            logger.info("Tracer file found, checking tracers.")
        # Get ppm, mass_accuracy_tr, yaxis_scale, and trendline_shown parameters
        ppm = self.parameters["mass_accuracy_units_tr"][1] == "ppm"
        mass_accuracy_tr = float(self.parameters["mass_accuracy_tr"][1])
        yaxis_scale = self.parameters["tracer_plot_yaxis_format"][1]
        trendline_shown = self.parameters["tracer_plot_trendline"][1] == "yes"
        # Perform check_tracers() on dfs with tracer file
        self.tracer_dfs_out = [
            (
                task_fun.check_feature_tracers(
                    df,
                    self.tracer_df,
                    mass_accuracy_tr,
                    float(self.parameters["rt_accuracy_tr"][1]),
                    ppm,
                    self.blank_headers,
                    self.sample_headers,
                )[0]
                if df is not None
                else None
            )
            for df in self.dfs
        ]
        self.dfs = [
            (
                task_fun.check_feature_tracers(
                    df,
                    self.tracer_df,
                    mass_accuracy_tr,
                    float(self.parameters["rt_accuracy_tr"][1]),
                    ppm,
                    self.blank_headers,
                    self.sample_headers,
                )[1]
                if df is not None
                else None
            )
            for df in self.dfs
        ]
        # Call format_tracer_file imported from utilities.py
        self.tracer_dfs_out = [
            task_fun.format_tracer_file(df) if df is not None else None for df in self.tracer_dfs_out
        ]

        # Declare plotter
        df_WA = WebApp_plotter()

        # Create plot
        if self.tracer_dfs_out[0] is not None:
            # Troubleshooting NTAW-460
            # logger.info("self.tracer_dfs_out[1] shape= {}".format(self.tracer_dfs_out[0].shape))
            # logger.info("self.tracer_dfs_out[1] columns= {}".format(self.tracer_dfs_out[0].columns.values))

            # Check that run sequence file is present. Raise IndexError if no run sequence file is found.
            if self.run_sequence_pos_df is None:
                raise IndexError(
                    "A run sequence file was not found. Please input a corresponding run sequence file for each input data file."
                )
            # If run sequence file present, check sample names in run sequence against those in the dataframe.
            else:
                # Pass positive mode data and positive run sequence file to check_run_seq()
                task_fun.check_run_seq(self.dfs[0], self.run_sequence_pos_df, self.blank_headers, self.sample_headers)

            listOfPNGs, df_debug, debug_list = df_WA.make_seq_scatter(
                df_in=self.tracer_dfs_out[0],
                df_seq=self.run_sequence_pos_df,
                ionization="pos",
                y_scale=yaxis_scale,
                fit=trendline_shown,
                share_y=False,
                # y_fixed=False,
                y_step=6,
                same_frame=False,
                legend=True,
                # chemical_names=None,
                dark_mode=False,
            )

            self.tracer_plots_out.append(listOfPNGs)

        else:
            self.tracer_plots_out.append(None)

        # Declare plotter
        df_WA = WebApp_plotter()

        # Create plot
        if self.tracer_dfs_out[1] is not None:
            # logger.info("self.tracer_dfs_out[1] shape= {}".format(self.tracer_dfs_out[1].shape))
            # logger.info("self.tracer_dfs_out[1] columns= {}".format(self.tracer_dfs_out[1].columns.values))

            # Check that run sequence file is present. Raise IndexError if no run sequence file is found.
            if self.run_sequence_neg_df is None:
                raise IndexError(
                    "A run sequence file was not found. Please input a corresponding run sequence file for each input data file."
                )
            # If run sequence file present, check sample names in run sequence against those in the dataframe.
            else:
                # Pass negative mode data and negative run sequence file to check_run_seq()
                task_fun.check_run_seq(self.dfs[1], self.run_sequence_neg_df, self.blank_headers, self.sample_headers)

            listOfPNGs, df_debug, debug_list = df_WA.make_seq_scatter(
                df_in=self.tracer_dfs_out[1],
                df_seq=self.run_sequence_neg_df,
                ionization="neg",
                y_scale=yaxis_scale,
                fit=trendline_shown,
                share_y=False,
                # y_fixed=False,
                y_step=6,
                same_frame=False,
                legend=True,
                # chemical_names=None,
                dark_mode=False,
            )

            self.tracer_plots_out.append(listOfPNGs)
            logger.info("df_debug shape= {}".format(df_debug.shape))
            logger.info("df_debug columns= {}".format(df_debug.columns.values))

        else:
            self.tracer_plots_out.append(None)

        # Combine items of tracer_dfs_out list
        dft = pd.concat([self.tracer_dfs_out[0], self.tracer_dfs_out[1]])
        # If passthrough columns present, combine
        if self.pass_through[0] is not None and self.pass_through[1] is not None:
            passthru = pd.concat([self.pass_through[0], self.pass_through[1]])
        elif self.pass_through[0] is not None:
            passthru = self.pass_through[0]
        else:
            passthru = self.pass_through[1]
        # Update logger
        logger.info("tracer dft= {}".format(dft.columns.tolist()))
        logger.info("tracer passthru= {}".format(passthru.columns.tolist()))
        # Combine and sort processed tracer file and combined passthrough columns
        self.data_map["Tracer Detection Statistics"] = task_fun.column_sort_TSR(dft, passthru)

        if self.tracer_plots_out[0] is not None:
            for i in range(len(self.tracer_plots_out[0])):
                self.tracer_map["tracer_plot_pos_" + str(i + 1)] = self.tracer_plots_out[0][i]

        # Add an if statement below to account for: if only negative mode data is entered, and only a negative tracer file is submitted, tracer_plots_out will only have one entry at [0]
        if len(self.tracer_plots_out) > 1:
            if self.tracer_plots_out[1] is not None:
                for i in range(len(self.tracer_plots_out[1])):
                    self.tracer_map["tracer_plot_neg_" + str(i + 1)] = self.tracer_plots_out[1][i]

        # Convert the figure objects in tracer_map into PNGs that can be stored in gridfs
        for key in self.tracer_map.keys():
            buf = io.BytesIO()
            # Save the figure in buffer as png
            self.tracer_map[key].savefig(buf, bbox_inches="tight", format="png")
            buf.seek(0)
            self.tracer_map[key] = buf.read()
        # Set project name
        project_name = self.parameters["project_name"][1]
        self.gridfs.put(
            "&&".join(self.tracer_map.keys()),
            _id=self.jobid + "_tracer_files",
            encoding="utf-8",
            project_name=project_name,
        )
        for key in self.tracer_map.keys():
            self.mongo_save(self.tracer_map[key], step=key)
        return

    def clean_features(self):
        """
        Accesses self.dfs and the user-input parameters for min replicate hits, max CV,
        and mrl std_dev multiplier. Check for presence of tracer file, then call task_fun.clean_features()
        that uses the parameters to document filtering and flagging the the Decision
        Documentation excel sheet. Call task_fun.Blank_Subtract_Mean() to subtract each feature's
        blank value from the mean of each sample grouping.

        Args:
            None
        Returns:
            None
        """
        # Define controls
        controls = [
            float(self.parameters["min_replicate_hits"][1]),
            float(self.parameters["max_replicate_cv"][1]),
            float(self.parameters["min_replicate_hits_blanks"][1]),
            float(self.parameters["mrl_std_multiplier"][1]),
        ]
        # Check for tracer file
        tracer_df_bool = False
        if self.tracer_df is not None:
            tracer_df_bool = True
        # Pass inputs to clean_features() and store docs, dfs_flagged, and finally dfs
        self.docs = [
            task_fun.clean_features(df, controls, tracer_df=tracer_df_bool)[1] if df is not None else None
            for index, df in enumerate(self.dfs)
        ]
        self.dfs_flagged = [
            task_fun.clean_features(df, controls, tracer_df=tracer_df_bool)[2] if df is not None else None
            for index, df in enumerate(self.dfs)
        ]
        self.dfs = [
            task_fun.clean_features(df, controls, tracer_df=tracer_df_bool)[0] if df is not None else None
            for index, df in enumerate(self.dfs)
        ]
        # subtract blanks from means
        self.dfs = [task_fun.Blank_Subtract_Mean(df) if df is not None else None for index, df in enumerate(self.dfs)]
        # subtract blanks from means
        self.dfs_flagged = [
            task_fun.Blank_Subtract_Mean(df) if df is not None else None for index, df in enumerate(self.dfs_flagged)
        ]
        return

    def merge_columns_onto_tracers(self):
        """
        Check for tracer files. If present, subset tracer dataframe with specific
        order of columns desired for the Tracer Summary output.

        Args:
            None
        Returns:
            None
        """
        # Check for tracer file, if not present return
        if self.tracer_df is None:
            logger.info("No tracer file, skipping this step.")
            return
        # Grab the tracer dataframe from self.data_map
        dft = self.data_map["Tracer Detection Statistics"]
        logger.info("dft: {}".format(dft.columns))
        # create summary table
        if "DTXSID" not in dft.columns:
            dft["DTXSID"] = ""
        # Subset dataframe with columns of interest
        dft = dft[
            [
                "Feature ID",
                "Chemical Name",
                "DTXSID",
                "Ionization Mode",
                "Mass Error (PPM)",
                "Retention Time Difference",
                "Total Detection Count",
                "Total Detection Percentage",
                "Max CV Across Samples",
            ]
        ]
        # Map dataframe to Tracer Summary output
        self.data_map["Tracer Summary"] = dft
        return

    def combine_modes(self):
        """
        For self.dfs and self.dfs_flagged, call task_fun.combine(). Checks for presence
        of pos and neg data modes, and if both present combine them. For self.doc_combined,
        call task_fun.combine_doc() function to check presence of both data modes and combine
        if present. Finally, merge passthrough columns back on to processed data (df_combined)
        and decision documentation sheet (doc_combined) by calling task_fun.MPP_ready().
        Store final dataframes in self.data_map dictionary.

        Args:
            None
        Returns:
            None
        """
        # Check for tracer file
        tracer_df_bool = False
        if self.tracer_df is not None:
            tracer_df_bool = True
        # combine dfs from both modes
        self.df_combined = task_fun.combine(self.dfs[0], self.dfs[1])
        # combine df_flaggeds from both modes
        self.df_flagged_combined = task_fun.combine(self.dfs_flagged[0], self.dfs_flagged[1])
        # combine docs from both modes
        self.doc_combined = task_fun.combine_doc(self.docs[0], self.docs[1], tracer_df=tracer_df_bool)

        # NTAW-577: Replace zero values of "Selected MRL" column with blank cells prior to exporting to data_map
        self.doc_combined["Selected MRL"] = self.doc_combined["Selected MRL"].replace(0, "")
        # Also replace zero values of blank mean columns with blank cells prior to exporting to data_map
        blanks = ["MB", "mb", "mB", "Mb", "blank", "Blank", "BLANK"]
        Mean = self.doc_combined.columns[self.doc_combined.columns.str.contains(pat="Mean ")].tolist()
        Mean_MB = [md for md in Mean if any(x in md for x in blanks)]
        self.doc_combined[Mean_MB] = self.doc_combined[Mean_MB].replace(0, "")
        # Map to Decision Documentation output
        self.data_map["Decision Documentation"] = self.doc_combined
        # Prep combined df for output by combining with passthrough cols and formatting
        self.mpp_ready = task_fun.MPP_Ready(
            self.df_combined,
            self.pass_through,
            self.blank_headers,
            self.sample_headers,
        )
        # Prep combined df_flagged for output by combining with passthrough cols and formatting
        self.mpp_ready_flagged = task_fun.MPP_Ready(
            self.df_flagged_combined,
            self.pass_through,
            self.blank_headers,
            self.sample_headers,
        )
        # Map df and df_flagged outputs to Final Occurrence Matrices sheets
        self.data_map["Final Occurrence Matrix"] = reduced_file(self.mpp_ready)
        self.data_map["Final Occurrence Matrix (flags)"] = reduced_file(self.mpp_ready_flagged)

    def perform_dashboard_search(self, lower_index=0, upper_index=None, save=True):
        """
        Access self.df_flagged_combined to get full list of chemical features that passed
        replication and MRL thresholds (flagged_combined keeps CV fails). Subset the dataframe
        by features flagged "For_Dashboard_Search". Check user-selected "search mode" - if "mass",
        call task_fun.masses() function to format masses for search. If "formulas", call
        task_fun.formulas() funcion to format formulas for search. Send formatted list to
        api_search_masses_batch() function imported from utilities.py. Handle returned JSON
        and combine back into results. Finally, call task_fun.calc_toxcast_percent_active()
        function to get tox information, and combine with dashboard results in "Chemical Results" output.

        Args:
            lower_index (integer; start search range at this Feature_ID)
            upper_index (interger, default=None; end search range at this Feature_ID)
        Returns:
            None
        """
        # Update logger
        logging.info(
            "Rows flagged for dashboard search: {} out of {}".format(
                len(self.df_flagged_combined.loc[self.df_flagged_combined["For_Dashboard_Search"] == "1", :]),
                len(self.df_flagged_combined),
            )
        )
        # Get subset of features to search from df_flagged_combined
        to_search = self.df_flagged_combined.loc[self.df_flagged_combined["For_Dashboard_Search"] == "1", :].copy()
        # Check if searching by mass or formula, drop duplicates
        if self.parameters["search_mode"][1] == "mass":
            to_search.drop_duplicates(subset="Mass", keep="first", inplace=True)
        else:
            to_search.drop_duplicates(subset="Formula", keep="first", inplace=True)
        # Calculate number of fragments to search
        n_search = len(to_search)
        # Update logger
        logger.info("Total # of queries: {}".format(n_search))
        # Subset features to search
        to_search = to_search.iloc[lower_index:upper_index, :]
        # If searching by mass, call api_search_masses_batch()
        if self.parameters["search_mode"][1] == "mass":
            mono_masses = task_fun.masses(to_search)
            dsstox_search_df = api_search_masses_batch(
                mono_masses,
                float(self.parameters["parent_ion_mass_accuracy"][1]),
                batchsize=150,
                jobid=self.jobid,
            )
        # If searching by formula, call api_search_masses_batch()
        else:
            formulas = task_fun.formulas(to_search)
            response = api_search_formulas(formulas, self.jobid)
            if not response.ok:  # check if we got a successful response
                raise requests.exceptions.HTTPError(
                    "Unable to access DSSTOX API. Please contact an administrator or try turning the DSSTox search option off."
                )
            dsstox_search_json = io.StringIO(json.dumps(response.json()["results"]))
            dsstox_search_df = pd.read_json(
                dsstox_search_json,
                orient="split",
                dtype={"TOXCAST_NUMBER_OF_ASSAYS/TOTAL": "object"},
            )
        # Merge API search results with mpp_ready_flagged dataframe ID, mass, and RT
        dsstox_search_df = self.mpp_ready_flagged[["Feature ID", "Mass", "Retention Time"]].merge(
            dsstox_search_df, how="right", left_on="Mass", right_on="INPUT"
        )
        # Calculate toxcast_percent_active values
        dsstox_search_df = task_fun.calc_toxcast_percent_active(dsstox_search_df)

        # Map dataframe to Chemical Results output
        self.data_map["Chemical Results"] = dsstox_search_df
        # Store search results
        self.search_results = dsstox_search_df

    def perform_hcd_search(self):
        """
        Check length of Dashboard Search Results. If length > 0, format Dashboard Search
        Results into list of unique DTXSIDs and call batch_search_hcd() function imported
        from utilities.py to retrieve hazard data from the Hazard Comparison Dashboard.
        Merge HCD results with Dashboard results and update "Chemical Results" output.

        Args:
            None
        Returns:
            None
        """
        # Update logger
        logger.info(f"Querying the HCD with DTXSID identifiers")
        # Check length of search results from API batch search
        if len(self.search_results) > 0:
            # Get unique DTXSIDs from searched features
            dtxsid_list = self.search_results["DTXSID"].unique()
            # Hit HCD API with unique DTXSIDs, retrieve hazard info
            hcd_results = batch_search_hcd(dtxsid_list)
            # Merge retrieved hazard info with dashboard search results
            self.search_results = self.search_results.merge(hcd_results, how="left", on="DTXSID")
            # Update Chemical Results output
            self.data_map["Chemical Results"] = self.search_results

    def mongo_save(self, file, step=""):
        """
        Take file chunk, and if it is a dataframe, arrange it via JSON to save in Mongo.
        Else, save separately.

        Args:
            file (any file type)
        Returns:
            None
        """
        # Check if file is type Pandas dataframe
        if isinstance(file, pd.DataFrame):
            # Convert to JSON
            to_save = file.to_json(orient="split")
        else:
            # Else handle with no conversion
            to_save = file
        # Format id
        id = self.jobid + "_" + step
        # Get project name
        project_name = self.parameters["project_name"][1]
        # Save item to MongoDB using item id
        self.gridfs.put(to_save, _id=id, encoding="utf-8", project_name=project_name)

    def save_excel_to_mongo(self):
        # Create an excel sheet from the datamap and save it to MongoDB
        in_memory_buffer = io.BytesIO()
        # Obtain a list of all keys in the data map (These will become the excel workbook sheet names)
        keys_list = list(self.data_map.keys())
        chemical_results_present = None
        if "Chemical Results" in keys_list:
            chemical_results_present = True
            # Replaces the static DTXSIDs in the DTXSID column with the corresponding hyperlinks.
            self.data_map["Chemical Results"]["DTXSID"] = self.data_map["Chemical Results"]["DTXSID"].apply(
                lambda x: make_hyperlink(x)
            )
            sheet_number = keys_list.index("Chemical Results")
        # Convert self.data_map dictionary into an excel workbook
        with pd.ExcelWriter(in_memory_buffer, engine="openpyxl") as writer:
            workbook = writer.book
            for df_name, df in self.data_map.items():
                df.to_excel(writer, sheet_name=df_name, index=False)
                # Format column widths to fit the largest string contained within the column
                sheet_num = keys_list.index(df_name)
                sheet = workbook.worksheets[sheet_num]
                # Freezes the top row of every sheet in the excel file.
                sheet.freeze_panes = "A2"
                # Format each column width to fit the longest string contained within the column
                for column in df:
                    try:
                        column_width = max(df[column].astype(str).map(len).max(), len(column)) + 1
                        col_idx = df.columns.get_loc(column) + 1
                        col_letter = get_column_letter(col_idx)
                        sheet.column_dimensions[col_letter].width = column_width
                    # NTAW-704: handle error where df[column] is recognised as a DataFrame, not a series
                    except AttributeError:
                        pass
            # Format DTXSID column hyperlinks an column width in the Chemical Results sheet
            if chemical_results_present:
                workbook = writer.book
                sheet = workbook.worksheets[sheet_number]
                for i in range(sheet.max_row):
                    cell = sheet.cell(row=i + 2, column=8)
                    cell.style = "Hyperlink"
                sheet.column_dimensions["H"].width = 18
        excel_data = in_memory_buffer.getvalue()

        # Save project name to MongoDB using jobid
        project_name = self.parameters["project_name"][1]
        self.gridfs.put(project_name, _id=f"{self.jobid}_project_name", encoding="utf-8")

        # Save results excel file to MongoDB using id
        id = self.jobid + "_excel"
        self.gridfs.put(excel_data, _id=id)
